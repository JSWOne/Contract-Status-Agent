# Tool Name   : Excel Tracker Updater via Power Automate
# Skill       : Contract Status Agent
# Version     : 1.0.0
# Last Updated: 2026-04-30
# Description : Takes a contract list, builds an in-memory .xlsx file using
#               openpyxl, base64-encodes the bytes, and POSTs
#               { "content": "<base64>" } to the Power Automate Excel-update flow.
# Dependencies: requests, openpyxl, python-dotenv
# ENV Vars    : PA_EXCEL_URL, DRY_RUN
# DRY_RUN     : Builds and encodes the Excel file but skips the POST; prints payload size

import argparse
import base64
import io
import json
import os
from datetime import datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


load_dotenv(Path(__file__).with_name(".env"))

DRY_RUN = os.getenv("DRY_RUN", "True").strip().lower() == "true"
PA_EXCEL_URL = os.getenv("PA_EXCEL_URL")

TOOL_NAME = "update_excel_via_pa.py"
SKILL_NAME = "Contract Status Agent"
MEMORY_PATH = Path(__file__).parent.parent / "Memory" / "memory.json"
LOG_PATH = Path("/tmp/error.log") if os.environ.get("GCS_MEMORY_BUCKET") else Path(__file__).parent.parent / "Logs" / "error.log"

HEADERS = ["Contract No", "Account Name", "Status", "Created Date"]


def log_error(
    run_id,
    step,
    error_code,
    error_message,
    payload=None,
    resolution=None,
    status="pending",
    ticket_id=None,
    learning=None,
):
    entry = {
        "timestamp": datetime.utcnow().isoformat(),
        "run_id": run_id,
        "skill": SKILL_NAME,
        "tool": TOOL_NAME,
        "step": step,
        "error_code": error_code,
        "error_message": error_message,
        "input_payload": payload or {},
        "context": step,
        "resolution_attempted": resolution,
        "resolution_status": status,
        "learning": learning,
        "ticket_id": ticket_id,
    }
    try:
        LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry) + "\n")
    except Exception as log_exc:
        import sys
        print(json.dumps(entry), file=sys.stderr)
        print(f"[log_error] Could not write to {LOG_PATH}: {log_exc}", file=sys.stderr)


def read_memory():
    if MEMORY_PATH.exists():
        return json.loads(MEMORY_PATH.read_text(encoding="utf-8"))
    return {
        "skill": SKILL_NAME,
        "last_run": None,
        "last_action": None,
        "state": {"run_history": []},
        "known_issues": [],
    }


def write_memory_step(memory, step_name, status, detail="", extra=None):
    step_entry = {
        "step": step_name,
        "status": status,
        "detail": detail,
        "timestamp": datetime.utcnow().isoformat(),
    }
    if extra:
        step_entry.update(extra)

    run_history = memory.setdefault("state", {}).setdefault("run_history", [])
    if run_history:
        run_history[-1].setdefault("steps", []).append(step_entry)

    memory["last_action"] = detail
    with open(MEMORY_PATH, "w", encoding="utf-8") as f:
        json.dump(memory, f, indent=2)


def normalize_contract(contract):
    return {
        "contract_no": str(contract.get("contract_no", "")).strip(),
        "account_name": str(contract.get("account_name", "")).strip(),
        "status": str(contract.get("status", "")).strip(),
        "created_date": str(contract.get("created_date", "")).strip(),
    }


def build_workbook_bytes(contracts):
    wb = Workbook()
    ws = wb.active
    ws.title = "LIVE Contracts"

    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for contract in contracts:
        normalized = normalize_contract(contract)
        ws.append(
            [
                normalized["contract_no"],
                normalized["account_name"],
                normalized["status"],
                normalized["created_date"],
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 16,
        "B": 36,
        "C": 24,
        "D": 24,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True)

    for idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(idx)].bestFit = True

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_payload(contracts):
    workbook_bytes = build_workbook_bytes(contracts)
    encoded = base64.b64encode(workbook_bytes).decode("utf-8")
    return {"content": encoded}, len(workbook_bytes)


def update_excel(contracts, run_id="run_manual", memory=None):
    if not isinstance(contracts, list):
        raise ValueError("contracts must be a list of dictionaries")

    memory = memory or read_memory()
    payload, workbook_size = build_payload(contracts)
    payload_size = len(payload["content"])

    if DRY_RUN:
        detail = (
            f"DRY_RUN: built Excel payload for {len(contracts)} contracts "
            f"({workbook_size} bytes workbook, {payload_size} chars base64)"
        )
        print(detail)
        write_memory_step(
            memory,
            "excel_update",
            "skipped",
            detail,
            extra={"excel_updated": False, "dry_run": True},
        )
        return {
            "excel_updated": False,
            "dry_run": True,
            "contracts_count": len(contracts),
            "workbook_bytes": workbook_size,
            "payload_chars": payload_size,
        }

    if not PA_EXCEL_URL:
        message = "PA_EXCEL_URL is missing from .env"
        log_error(run_id, "excel_update", "PA_CONFIG_MISSING", message, status="failed")
        write_memory_step(memory, "excel_update", "failed", message, extra={"excel_updated": False})
        raise RuntimeError(message)

    try:
        response = requests.post(PA_EXCEL_URL, json=payload, timeout=60)
        if 400 <= response.status_code < 500:
            raise requests.HTTPError(f"PA_4XX {response.status_code}: {response.text}")
        if response.status_code >= 500:
            raise requests.HTTPError(f"PA_500 {response.status_code}: {response.text}")
        response.raise_for_status()

        if response.status_code == 202:
            detail = (
                f"Power Automate flow accepted Excel update request for {len(contracts)} contracts; "
                "check flow run history for final update status"
            )
            excel_updated = None
        else:
            detail = f"Excel updated via Power Automate for {len(contracts)} contracts"
            excel_updated = True

        write_memory_step(
            memory,
            "excel_update",
            "success",
            detail,
            extra={
                "excel_updated": excel_updated,
                "flow_triggered": True,
                "status_code": response.status_code,
            },
        )
        return {
            "excel_updated": excel_updated,
            "flow_triggered": True,
            "dry_run": False,
            "contracts_count": len(contracts),
            "status_code": response.status_code,
            "note": "HTTP 202 means the flow was triggered; confirm final success in Power Automate run history.",
        }

    except requests.HTTPError as e:
        error_code = "PA_500" if "PA_500" in str(e) else "PA_4XX"
        log_error(
            run_id,
            "excel_update",
            error_code,
            str(e),
            payload={"contracts_count": len(contracts), "payload_chars": payload_size},
            status="failed",
        )
        write_memory_step(memory, "excel_update", "failed", str(e), extra={"excel_updated": False})
        raise

    except requests.RequestException as e:
        log_error(
            run_id,
            "excel_update",
            "PA_REQUEST_FAIL",
            str(e),
            payload={"contracts_count": len(contracts), "payload_chars": payload_size},
            resolution="Check PA_EXCEL_URL and network connectivity; retry once from orchestrator",
            status="pending",
        )
        write_memory_step(memory, "excel_update", "failed", str(e), extra={"excel_updated": False})
        raise


def load_contracts_from_json(path):
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if isinstance(data, dict) and "contracts" in data:
        data = data["contracts"]
    if not isinstance(data, list):
        raise ValueError("JSON input must be a list, or an object with a 'contracts' list")
    return data


def main():
    parser = argparse.ArgumentParser(description="Update Contract Status Excel tracker via Power Automate.")
    parser.add_argument("--contracts-json", required=True, help="Path to JSON file containing scraped contracts.")
    parser.add_argument("--run-id", default="run_manual", help="Run id for memory/error logging.")
    args = parser.parse_args()

    contracts = load_contracts_from_json(args.contracts_json)
    result = update_excel(contracts, run_id=args.run_id)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
