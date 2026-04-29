"""
Tool: salesforce_contract_monitor.py
Purpose: Monitor Salesforce JSW One All Contracts list for status changes.

Logic (every 15 minutes):
  1. Scrape ALL contracts from the list view (no status filter).
  2. Merge into the master file (.tmp/contracts_master.json) —
     adds new contracts, updates existing ones.
  3. Filter master file: only contracts created within the last 30 days.
  4. Compare filtered contracts' current status vs previous status.
  5. Alert (Teams Adaptive Card) for any status change.

Session files (.tmp/):
    sf_session.json        — Playwright storage state (cookies / localStorage)
    contracts_master.json  — Master record of all scraped contracts
"""

import os
import csv
import json
import base64
import logging
from datetime import datetime, timedelta

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from playwright.sync_api import sync_playwright
from dotenv import load_dotenv

load_dotenv()

log = logging.getLogger(__name__)

BASE_URL      = "https://jswsteel.my.site.com"
CONTRACTS_URL = BASE_URL + "/jswone/s/recordlist/Contract/Default?Contract-filterId=JSW_One_All_Contracts"

_BASE_DIR      = os.path.join(os.path.dirname(__file__), "..")
SESSION_FILE   = os.path.join(_BASE_DIR, ".tmp", "sf_session.json")
MASTER_FILE    = os.path.join(_BASE_DIR, ".tmp", "contracts_master.json")
SCRAPED_CSV    = os.path.join(_BASE_DIR, ".tmp", "contracts_scraped.csv")
APPROVAL_CSV   = os.path.join(_BASE_DIR, ".tmp", "contracts_approval_master.csv")
TRACKER_EXCEL       = os.path.join(_BASE_DIR, ".tmp", "JSW ONE Agent SO Live Tracker.xlsx")
ONEDRIVE_EXCEL_PATH = os.environ.get("ONEDRIVE_EXCEL_PATH", "")
DEBUG_DIR      = os.path.join(_BASE_DIR, ".tmp")

# Module-level singletons — all Playwright calls must happen on the monitor thread
_pw      = None
_browser = None
_context = None
_page    = None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def initialize_session() -> None:
    """
    Launch headless Chromium, restore saved session if available,
    navigate to the contracts page, and login if redirected.
    Call once from the monitor thread before starting the poll loop.
    """
    global _pw, _browser, _context, _page

    # Clean up any previous Playwright instance — its internal event loop
    # must be stopped before starting a new one or Playwright raises asyncio conflict.
    if _pw is not None:
        try:
            _pw.stop()
        except Exception:
            pass
        _pw = None

    os.makedirs(DEBUG_DIR, exist_ok=True)
    log.info("[monitor] Initialising Salesforce session…")

    _pw      = sync_playwright().start()
    _browser = _pw.chromium.launch(
        headless=True,
        args=[
            "--no-sandbox",           # required in Cloud Run / Docker containers
            "--disable-dev-shm-usage", # /dev/shm is only 64MB in Cloud Run; use /tmp
            "--disable-gpu",
        ],
    )

    # Mimic a real desktop Chrome browser — headless Chromium from a datacenter
    # is detected by Salesforce and served a bot-challenge / blank page otherwise.
    _ctx_args = dict(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/131.0.0.0 Safari/537.36"
        ),
        viewport={"width": 1920, "height": 1080},
        locale="en-US",
    )

    if os.path.exists(SESSION_FILE):
        log.info("[monitor] Restoring session from %s", SESSION_FILE)
        try:
            _context = _browser.new_context(storage_state=SESSION_FILE, **_ctx_args)
        except Exception as e:
            log.warning("[monitor] Session restore failed (%s) — fresh context", e)
            _context = _browser.new_context(**_ctx_args)
    else:
        _context = _browser.new_context(**_ctx_args)

    _page = _context.new_page()
    _navigate_and_login_if_needed()
    log.info("[monitor] Session ready. URL: %s", _page.url)


def check_for_changes() -> list:
    """
    Main poll function called every 15 minutes:
      1. Scrape all contracts from Salesforce (sorted by Created Date desc).
      2. Filter to contracts currently in 'In Approval Process'.
      3. Add new 'In Approval Process' contracts to master file.
      4. Detect status changes for any contract previously tracked in master.
      5. Return list of status changes.

    Each change dict: {contract_no, account_name, old_status, new_status, url, created_date}
    """
    is_first_run = not os.path.exists(MASTER_FILE)

    master  = _load_master()
    current = scrape_all_contracts()   # also saves contracts_scraped.csv internally

    # Contracts (from last 2 weeks) currently in "In Approval Process"
    in_approval = {
        no: data for no, data in current.items()
        if data.get("status", "").strip() == "In Approval Process"
    }

    # ---- First run: print all In Approval Process contracts ----
    if is_first_run:
        log.info("[monitor] ── FIRST RUN ── Scanned last 2 weeks. Found %d contracts in 'In Approval Process':",
                 len(in_approval))
        for no, d in sorted(in_approval.items()):
            log.info("[monitor]   %s | %-45s | Created: %s",
                     no, d["account_name"], d["created_date"])

    # ---- Subsequent runs: summary + new contracts entering approval ----
    else:
        new_contracts = [no for no in in_approval if no not in master]
        log.info("[monitor] ── CHECKING UPDATES ── Scanned last 2 weeks. %d contracts in 'In Approval Process' (%d tracked in master)",
                 len(in_approval), len(master))
        if new_contracts:
            log.info("[monitor] NEW contracts entered 'In Approval Process' (%d):", len(new_contracts))
            for no in new_contracts:
                d = in_approval[no]
                log.info("[monitor]   + %s | %s | Created: %s",
                         no, d["account_name"], d["created_date"])
        else:
            log.info("[monitor] No new 'In Approval Process' contracts.")

    # ---- Detect status changes for previously tracked contracts ----
    changes = []
    for contract_no, master_data in master.items():
        current_data = current.get(contract_no)
        if not current_data:
            continue  # contract no longer visible in list view
        old_status = master_data["status"]
        new_status = current_data["status"]
        if old_status != new_status:
            changes.append({
                "contract_no":  contract_no,
                "account_name": current_data["account_name"],
                "old_status":   old_status,
                "new_status":   new_status,
                "url":          current_data["url"],
                "created_date": current_data.get("created_date", ""),
            })
            log.info("[monitor] STATUS CHANGE: %s  %s → %s",
                     contract_no, old_status, new_status)

    if not changes:
        log.info("[monitor] No status changes detected.")

    # ---- Update master ----
    # Add new In Approval Process contracts
    for no, data in in_approval.items():
        if no not in master:
            master[no] = data.copy()
        else:
            master[no].update(data)  # refresh url / account_name

    # Apply status updates for changed contracts
    for change in changes:
        no = change["contract_no"]
        if no in master:
            master[no]["status"] = change["new_status"]

    _save_master(master)
    log.info("[monitor] Master file updated: %d tracked contracts.", len(master))

    return changes


def scrape_all_contracts() -> dict:
    """
    Navigate to JSW One All Contracts, sort by Created Date descending,
    load records only up to 2 weeks back, and return those contracts as
    {contract_no: {status, account_name, created_date, url}}.
    """
    _ensure_session_valid()

    log.info("[monitor] Navigating to contracts list…")
    _page.goto(CONTRACTS_URL, wait_until="commit", timeout=120_000)
    _page.wait_for_timeout(5_000)

    # Wait for the table to actually populate (Salesforce LWC can be slow after idle)
    try:
        _page.wait_for_selector("table tbody tr", timeout=60_000)
    except Exception:
        current_url = _page.url or ""
        if "/login" in current_url.lower():
            log.warning("[monitor] Session expired during navigation — re-authenticating")
            _do_login()
            _context.storage_state(path=SESSION_FILE)
            _page.goto(CONTRACTS_URL, wait_until="commit", timeout=120_000)
            _page.wait_for_timeout(8_000)
        else:
            log.warning("[monitor] Contracts table not visible after 65s (URL: %s) — continuing", current_url)
            _screenshot("monitor_table_empty")

    _screenshot("monitor_list_loaded")

    _sort_created_date_desc()

    cutoff = datetime.now() - timedelta(days=14)
    _load_all_records(date_cutoff=cutoff)   # early-stop scans cells, no index needed
    _screenshot("monitor_all_loaded")

    col_map   = _get_column_map()           # headers stable after full load
    contracts = _extract_rows(col_map)      # re-detects indices inside JS

    # Belt-and-suspenders: drop anything older than the cutoff
    contracts = {
        no: data for no, data in contracts.items()
        if _parse_date(data.get("created_date", "")) is not None
        and _parse_date(data.get("created_date", "")) >= cutoff
    }

    # Save Excel tracker (all contracts → split into per-status sheets)
    _save_excel_tracker(contracts)

    log.info("[monitor] Scraped %d contracts within last 2 weeks (col_map=%s)", len(contracts), col_map)
    return contracts


# ---------------------------------------------------------------------------
# Session management
# ---------------------------------------------------------------------------

def _navigate_and_login_if_needed() -> None:
    """Check session validity via login URL; authenticate if needed.
    Does NOT navigate to the contracts page — scrape_all_contracts() does that once."""
    _validate_required_env()

    _page.goto(os.environ["SALESFORCE_URL"], wait_until="commit", timeout=120_000)
    _page.wait_for_timeout(8_000)

    if "/login" in _page.url.lower():
        # If we loaded a session file but still landed on the login page, the
        # cookies are expired. Salesforce renders a different auth variant with
        # those stale cookies (no standard Username input). Delete the file and
        # reload so the clean standard login form appears.
        if os.path.exists(SESSION_FILE):
            os.remove(SESSION_FILE)
            log.info("[monitor] Stale session detected — cleared %s, reloading login page", SESSION_FILE)
            _page.reload(wait_until="commit", timeout=120_000)
            _page.wait_for_timeout(5_000)

        log.info("[monitor] On login page — authenticating…")
        _do_login()
        _context.storage_state(path=SESSION_FILE)
        log.info("[monitor] Session saved to %s", SESSION_FILE)
    else:
        log.info("[monitor] Session still valid. URL: %s", _page.url)


def _do_login() -> None:
    """Fill and submit the Salesforce login form. Page is already on login URL."""
    # wait_until="commit" only guarantees the HTTP response started — the login
    # form is rendered by JavaScript which can take 2-3 min to download and run
    # on a cold GCP connection. 300s gives enough headroom for any realistic case.
    username = _first_visible_locator([
        'input[placeholder="Username"]',
        'input[placeholder*="Username"]',
        'input[placeholder*="Email"]',
        'input[name="username"]',
        'input[name="UserName"]',
        'input#username',
        'input[type="email"]',
        'input[autocomplete="username"]',
        'input[aria-label*="Username"]',
        'input[aria-label*="Email"]',
        'input:not([type="hidden"]):not([type="password"])',
    ], timeout=300_000)
    if username is None:
        _raise_login_form_error("username field")

    password = _first_visible_locator([
        'input[type="password"]',
        'input[name="pw"]',
        'input[name="password"]',
        'input#password',
        'input[autocomplete="current-password"]',
        'input[aria-label*="Password"]',
    ], timeout=60_000)
    if password is None:
        _raise_login_form_error("password field")

    _type_login_value(username, os.environ["SALESFORCE_USERNAME"], "username")
    _type_login_value(password, os.environ["SALESFORCE_PASSWORD"], "password")

    # Try button selectors in order — Salesforce sometimes renders "Log in" or "Login"
    button = _first_visible_locator([
        'button:has-text("Log in")',
        'button:has-text("Login")',
        'button[type="submit"]',
        'input[type="submit"]',
        '[role="button"]:has-text("Log in")',
        '[role="button"]:has-text("Login")',
    ], timeout=15_000)
    if button is None:
        _raise_login_form_error("login button")
    _submit_login(button, password)

    _screenshot("monitor_after_login_click")
    _page.wait_for_timeout(3_000)  # allow redirect chain to start

    try:
        _page.wait_for_url(lambda u: "/login" not in u.lower(), timeout=120_000)
    except Exception:
        _screenshot("monitor_login_failed")
        raise RuntimeError(
            f"Login timed out — still on login page. "
            f"Check credentials or MFA. URL: {_page.url}"
        )

    log.info("[monitor] Login successful. URL: %s", _page.url)
    _screenshot("monitor_logged_in")


def _submit_login(button, password) -> None:
    """Submit Salesforce login with fallbacks for headless Cloud Run clicks."""
    attempts = [
        ("keyboard Enter", lambda: _page.keyboard.press("Enter")),
        ("password Enter", lambda: password.press("Enter", timeout=5_000)),
        ("coordinate click", lambda: _coordinate_click(button)),
        ("normal click", lambda: button.click(timeout=5_000)),
        ("forced click", lambda: button.click(timeout=5_000, force=True)),
        ("DOM click", lambda: button.evaluate("(el) => el.click()")),
    ]

    last_error = None
    for label, action in attempts:
        try:
            log.info("[monitor] Submitting login via %s", label)
            action()
            if _wait_for_login_redirect(timeout=15_000):
                log.info("[monitor] Login submit succeeded via %s", label)
                return
        except Exception as exc:
            last_error = exc
            log.warning("[monitor] Login submit via %s failed: %s", label, exc)

    _raise_login_form_error(f"login submit action; last error: {last_error}")


def _type_login_value(locator, value: str, label: str) -> None:
    """Type into Aura login inputs so Salesforce component state is updated."""
    locator.click(timeout=10_000)
    locator.press("Control+A", timeout=5_000)
    locator.press("Backspace", timeout=5_000)
    locator.type(value, delay=30, timeout=60_000)
    log.info("[monitor] Typed %s field", label)


def _coordinate_click(locator) -> None:
    box = locator.bounding_box(timeout=5_000)
    if not box:
        raise RuntimeError("login button has no bounding box")
    _page.mouse.click(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)


def _wait_for_login_redirect(timeout: int) -> bool:
    try:
        _page.wait_for_url(lambda url: "/login" not in url.lower(), timeout=timeout)
        return True
    except Exception:
        return "/login" not in (_page.url or "").lower()


def _validate_required_env() -> None:
    missing = [
        name for name in (
            "SALESFORCE_URL",
            "SALESFORCE_USERNAME",
            "SALESFORCE_PASSWORD",
        )
        if not os.environ.get(name)
    ]
    if missing:
        raise RuntimeError(
            "Missing required environment variable(s): "
            + ", ".join(missing)
            + ". Set them on the Cloud Run service before starting the monitor."
        )


def _first_visible_locator(selectors: list, timeout: int):
    """Return the first visible locator from a selector list, waiting up to timeout."""
    deadline = datetime.now() + timedelta(milliseconds=timeout)
    last_error = None

    while datetime.now() < deadline:
        for selector in selectors:
            try:
                locator = _page.locator(selector).first
                if locator.is_visible(timeout=1_000):
                    log.info("[monitor] Login selector matched: %s", selector)
                    return locator
            except Exception as exc:
                last_error = exc
                continue
        _page.wait_for_timeout(1_000)

    if last_error:
        log.warning("[monitor] Last selector lookup error: %s", last_error)
    return None


def _raise_login_form_error(missing_part: str) -> None:
    """Capture enough page state to diagnose Cloud Run/Salesforce login issues."""
    safe_part = "".join(ch if ch.isalnum() else "_" for ch in missing_part)[:80]
    _screenshot(f"monitor_login_missing_{safe_part}")
    try:
        title = _page.title()
    except Exception:
        title = ""
    try:
        body_text = (_page.evaluate("() => document.body.innerText") or "").strip()[:1200]
    except Exception:
        body_text = ""
    try:
        inputs = _page.evaluate("""
            () => Array.from(document.querySelectorAll('input, button')).slice(0, 30).map(el => ({
                tag: el.tagName,
                type: el.getAttribute('type') || '',
                id: el.id || '',
                name: el.getAttribute('name') || '',
                placeholder: el.getAttribute('placeholder') || '',
                aria: el.getAttribute('aria-label') || '',
                text: (el.innerText || el.value || '').slice(0, 80),
            }))
        """)
    except Exception:
        inputs = []

    log.error("[monitor] Login page diagnostics: url=%s title=%s inputs=%s body=%r",
              _page.url, title, inputs, body_text)
    raise RuntimeError(
        f"Salesforce login page loaded, but the {missing_part} was not visible. "
        f"URL: {_page.url}. Check Cloud Run logs for login page diagnostics; "
        "Salesforce may be showing a bot challenge, SSO/MFA page, or a different login template."
    )


def _ensure_session_valid() -> None:
    """Re-initialise Playwright if the page is closed or session expired."""
    global _pw, _browser, _context, _page

    try:
        if _page is not None and not _page.is_closed():
            if "/login" in _page.url.lower():
                log.warning("[monitor] Session expired — re-authenticating")
                _navigate_and_login_if_needed()
            return
    except Exception:
        pass

    log.warning("[monitor] Playwright session lost — reinitialising")
    try:
        _pw.stop()
    except Exception:
        pass
    initialize_session()


# ---------------------------------------------------------------------------
# Scraping helpers
# ---------------------------------------------------------------------------

def _sort_created_date_desc() -> None:
    """
    Click the 'Created Date' column header until the list is sorted descending.
    Verifies success by checking that the first visible row has a recent date,
    rather than relying on aria-sort (which Salesforce LWC does not always update).
    """

    def _first_row_date() -> str:
        """Return the Created Date text of the first table row, or ''."""
        try:
            return _page.evaluate("""
                () => {
                    const dateRe = /\\d{2}\\/\\d{2}\\/\\d{4}/;
                    const rows = document.querySelectorAll('table tbody tr');
                    if (!rows.length) return '';
                    const cells = Array.from(rows[0].querySelectorAll('th, td'));
                    for (let i = cells.length - 1; i >= 0; i--) {
                        const t = (cells[i].innerText || cells[i].textContent || '').trim();
                        if (dateRe.test(t)) return t;
                    }
                    return '';
                }
            """)
        except Exception:
            return ""

    def _is_recent(date_str: str) -> bool:
        d = _parse_date(date_str)
        return d is not None and d >= datetime.now() - timedelta(days=30)

    try:
        # Find the Created Date column header
        target_th = None
        headers = _page.locator("table thead th, table thead td").all()
        for th in headers:
            try:
                text = th.inner_text(timeout=2_000).strip().lower()
                if "created" in text and "date" in text:
                    target_th = th
                    break
            except Exception:
                continue

        if target_th is None:
            log.warning("[monitor] Created Date column header not found — cannot sort")
            return

        # Already sorted descending?
        first_date = _first_row_date()
        if _is_recent(first_date):
            log.info("[monitor] Created Date already sorted descending (first row: %s)", first_date)
            return

        # Click up to 3 times, waiting for the table to reload each time
        for attempt in range(1, 4):
            target_th.scroll_into_view_if_needed()
            target_th.click(timeout=5_000)
            # Wait for Salesforce to reload the list (network + settle)
            try:
                _page.wait_for_load_state("domcontentloaded", timeout=10_000)
            except Exception:
                pass
            _page.wait_for_timeout(4_000)

            first_date = _first_row_date()
            log.info("[monitor] Sort attempt %d: first row date = '%s'", attempt, first_date)

            if _is_recent(first_date):
                log.info("[monitor] Sorted by Created Date descending after %d click(s)", attempt)
                _screenshot("monitor_sorted_desc")
                return

        log.warning("[monitor] Could not confirm descending sort after 3 attempts — proceeding anyway")
        _screenshot("monitor_sort_failed")

    except Exception as e:
        log.warning("[monitor] Could not sort by Created Date: %s — continuing", e)


def _load_all_records(date_cutoff=None) -> None:
    """
    Scroll down to trigger Salesforce's infinite-scroll and load all records
    within date_cutoff. Uses wait_for_function (polling) instead of a fixed
    sleep so we wait exactly as long as the XHR takes — no more, no less.
    """
    MAX_ITER = 100

    for i in range(MAX_ITER):
        current_row_count = _page.locator("table tbody tr").count()
        log.info("[monitor] Row count (iter %d): %d", i + 1, current_row_count)

        # Early-stop: last row older than 2-week cutoff
        if date_cutoff is not None and current_row_count > 0:
            try:
                last_date_str = _page.evaluate("""
                    () => {
                        const rows = document.querySelectorAll('table tbody tr');
                        const last = rows[rows.length - 1];
                        if (!last) return '';
                        const dateRe = /\\d{2}\\/\\d{2}\\/\\d{4}/;
                        for (const cell of Array.from(last.querySelectorAll('td')).reverse()) {
                            const t = (cell.textContent || '').trim();
                            if (dateRe.test(t)) return t;
                        }
                        return '';
                    }
                """)
                last_date = _parse_date(last_date_str)
                if last_date and last_date < date_cutoff:
                    log.info("[monitor] Reached 2-week cutoff at '%s' — %d rows loaded",
                             last_date_str, current_row_count)
                    _screenshot("monitor_load_done")
                    break
            except Exception:
                pass

        # Scroll to trigger the infinite-scroll XHR
        _page.evaluate("""
            () => {
                window.scrollTo(0, document.body.scrollHeight);
                const el = document.querySelector('.slds-scrollable_y, .forceListViewManagerBody');
                if (el) el.scrollTop = el.scrollHeight;
            }
        """)

        # Wait until new rows actually appear — handles any XHR latency automatically
        try:
            _page.wait_for_function(
                f"() => document.querySelectorAll('table tbody tr').length > {current_row_count}",
                timeout=120_000,
            )
        except Exception:
            # No new rows after 120s → all records are loaded
            log.info("[monitor] All records loaded: %d rows (%d iterations)",
                     current_row_count, i + 1)
            _screenshot("monitor_load_done")
            break


def _get_column_map() -> dict:
    """
    Read table headers and return semantic → column-index mapping.
    Defaults based on observed Salesforce layout:
      contract_no=6, account_name=2, status=3, created_date=7
    """
    col_map = {"contract_no": 6, "account_name": 2, "status": 3, "created_date": 7}
    try:
        headers = _page.locator("table thead th, table thead td").all()
        for i, th in enumerate(headers):
            try:
                text = th.inner_text(timeout=2_000).strip().lower()
                if "status" in text:
                    col_map["status"] = i
                elif "account" in text:
                    col_map["account_name"] = i
                elif "created" in text and "date" in text:
                    col_map["created_date"] = i
                elif "contract" in text and "date" not in text and "start" not in text \
                        and "end" not in text and "status" not in text:
                    col_map["contract_no"] = i
            except Exception:
                continue
    except Exception as e:
        log.warning("[monitor] Could not parse headers: %s — using defaults", e)
    return col_map


def _cell_text(cell) -> str:
    """Helper — not used at runtime; documents the JS helper approach."""
    pass


def _extract_rows(col_map: dict) -> dict:
    """
    Two-phase extraction:
      Phase 1 — save a raw CSV with every column exactly as shown on the website.
      Phase 2 — return {contract_no: {status, account_name, created_date, url}}.

    Contract number strategy (most robust):
      For each row, scan every <a> link's textContent for a 5+ digit run.
      textContent works even when innerText returns empty for LWC-rendered cells.
    """
    # ── Phase 1: raw all-columns CSV (debug) ─────────────────────────────────
    raw = _page.evaluate("""
        () => {
            const ths = Array.from(
                document.querySelectorAll('table thead th, table thead td')
            );
            const headers = ths.map(th =>
                (th.innerText || th.textContent || '').trim()
            );
            const rows = Array.from(document.querySelectorAll('table tbody tr'));
            const data = rows.map(row =>
                Array.from(row.querySelectorAll('th, td')).map(cell =>
                    (cell.innerText || cell.textContent || '').trim().replace(/\\n/g, ' ')
                )
            );
            return { headers, data };
        }
    """)
    _save_raw_all_columns_csv(raw.get("headers", []), raw.get("data", []))
    log.info("[monitor] Phase 1 raw CSV saved (%d rows). Sample first row: %s",
             len(raw.get("data", [])),
             raw.get("data", [[]])[0][:6] if raw.get("data") else [])

    # ── Phase 2: structured extraction with shadow DOM piercing ─────────────
    acc_default = col_map.get("account_name", 2)
    st_default  = col_map.get("status",       3)
    cd_default  = col_map.get("created_date", 7)

    result = _page.evaluate(
        """
        (defaults) => {
            const { accDef, stDef, cdDef } = defaults;

            // ── Shadow DOM text extractor ──────────────────────────────────
            // Salesforce LWC renders some cells (e.g. contract number) inside
            // an open shadow root.  Regular textContent stops at shadow
            // boundaries, so we must walk them explicitly.
            function nodeText(n) {
                if (!n) return '';
                if (n.nodeType === 3) return n.textContent || '';   // text node
                if (n.nodeType === 1) return shadowText(n);         // element
                return '';
            }
            function shadowText(el) {
                if (!el) return '';
                // Pierce shadow root first (open shadow in LWC)
                if (el.shadowRoot) {
                    let t = '';
                    for (const c of el.shadowRoot.childNodes) t += nodeText(c);
                    if (t.trim()) return t.trim();
                }
                // Regular children
                let t = '';
                for (const c of el.childNodes) t += nodeText(c);
                return t.trim();
            }
            // Return first-line text: prefer visible innerText, fall back to
            // shadow DOM traversal when innerText is empty (LWC shadow cells).
            function cellText(el) {
                if (!el) return '';
                const it = (el.innerText || '').trim();
                if (it) return it.split('\\n')[0].trim();
                const st = shadowText(el);
                return st ? st.split('\\n')[0].trim() : '';
            }

            // ── Detect column indices from headers ────────────────────────
            // Use FULL innerText (not split on newline) so multi-line headers
            // like "Sort\\nContract Number\\nShow..." still contain the keywords.
            const ths = Array.from(
                document.querySelectorAll('table thead th, table thead td')
            );
            const hdrTexts = ths.map(th =>
                (th.innerText || th.textContent || '').trim().toLowerCase()
            );
            let hdrCdIdx = -1, hdrStIdx = -1, hdrAccIdx = -1;
            hdrTexts.forEach((t, i) => {
                if      (t.includes('created') && t.includes('date')) hdrCdIdx  = i;
                else if (t.includes('status'))                         hdrStIdx  = i;
                else if (t.includes('account'))                        hdrAccIdx = i;
            });

            // ── Compute body/header offset ────────────────────────────────
            // Find actual Created Date td by matching "DD/MM/YYYY, H" pattern
            const dtRe     = /\\d{2}\\/\\d{2}\\/\\d{4},\\s*\\d/;
            const firstRow = document.querySelector('table tbody tr');
            const fCells   = firstRow ? Array.from(firstRow.querySelectorAll('td')) : [];
            let actualCdIdx = -1;
            for (let i = fCells.length - 1; i >= 0; i--) {
                if (dtRe.test(cellText(fCells[i]))) { actualCdIdx = i; break; }
            }
            const offset = (hdrCdIdx >= 0 && actualCdIdx >= 0)
                           ? actualCdIdx - hdrCdIdx : 0;

            const cdIdx  = actualCdIdx >= 0 ? actualCdIdx
                         : (hdrCdIdx  >= 0 ? hdrCdIdx  + offset : cdDef);
            const stIdx  = hdrStIdx  >= 0 ? hdrStIdx  + offset : stDef;
            const accIdx = hdrAccIdx >= 0 ? hdrAccIdx + offset : accDef;

            // ── Extract rows ──────────────────────────────────────────────
            const cnRe = /\\b(\\d{5,})\\b/;

            const rows = Array.from(document.querySelectorAll('table tbody tr'));
            const data = rows.map(row => {
                const cells = Array.from(row.querySelectorAll('td'));

                // Contract number lives in a <th scope="row"> element — Salesforce
                // renders the primary-key column as a row-header <th>, not a <td>.
                // Regular querySelectorAll('td') misses it entirely.
                let contractNo  = '';
                let contractUrl = '';
                const th = row.querySelector('th');
                if (th) {
                    const t = cellText(th);
                    const m = t.match(cnRe);
                    if (m) {
                        contractNo = m[1];
                        const a = th.querySelector('a');
                        if (a) contractUrl = a.href;
                    }
                }
                // Fallback: scan <td> cells (in case layout differs)
                if (!contractNo) {
                    for (const cell of cells) {
                        const t = cellText(cell);
                        const m = t.match(cnRe);
                        if (m) {
                            contractNo = m[1];
                            const a = cell.querySelector('a');
                            if (a) contractUrl = a.href;
                            break;
                        }
                    }
                }

                return {
                    contract_no:  contractNo,
                    account_name: cellText(cells[accIdx]),
                    status:       cellText(cells[stIdx]),
                    created_date: cellText(cells[cdIdx]),
                    url:          contractUrl,
                };
            });

            return { stIdx, accIdx, cdIdx, offset, rows: data };
        }
        """,
        {"accDef": acc_default, "stDef": st_default, "cdDef": cd_default}
    )

    rows_data = result.get("rows", [])
    log.info("[monitor] Column indices — account=%s, status=%s, created_date=%s (offset=%s)",
             result.get("accIdx"), result.get("stIdx"),
             result.get("cdIdx"), result.get("offset"))
    log.info("[monitor] Total rows in table: %d", len(rows_data))

    if rows_data:
        first = rows_data[0]
        log.info("[monitor] First row — contract_no='%s', account='%s', status='%s', created_date='%s'",
                 first.get("contract_no"), first.get("account_name"),
                 first.get("status"), first.get("created_date"))

    contracts = {}
    for row in rows_data:
        no = (row.get("contract_no") or "").strip()
        if not no:
            continue
        contracts[no] = {
            "status":       (row.get("status")       or "").strip(),
            "account_name": (row.get("account_name") or "").strip(),
            "created_date": (row.get("created_date") or "").strip(),
            "url":          row.get("url") or "",
        }

    return contracts


# ---------------------------------------------------------------------------
# Excel tracker
# ---------------------------------------------------------------------------

def _save_excel_tracker(contracts: dict) -> None:
    """
    Write (or overwrite) the Excel tracker with a single 'All SO' sheet
    containing all contracts from the last 14 days, sorted newest-first.
    Columns: Contract No | Account Name | Status | Created Date
    """
    COLS    = ["Contract No", "Account Name", "Status", "Created Date"]
    HDR_BG  = PatternFill("solid", fgColor="BDD7EE")
    HDR_FNT = Font(bold=True)

    sorted_rows = sorted(
        contracts.items(),
        key=lambda x: _parse_date(x[1].get("created_date", "")) or datetime.min,
        reverse=True,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LIVE Contracts"

    ws.append(COLS)
    for cell in ws[1]:
        cell.font      = HDR_FNT
        cell.fill      = HDR_BG
        cell.alignment = Alignment(horizontal="center")

    for no, data in sorted_rows:
        ws.append([no,
                   data.get("account_name", ""),
                   data.get("status", ""),
                   data.get("created_date", "")])

    col_widths = [len(c) + 2 for c in COLS]
    for no, data in sorted_rows:
        row = [no, data.get("account_name",""), data.get("status",""), data.get("created_date","")]
        for i, val in enumerate(row):
            col_widths[i] = max(col_widths[i], len(str(val)) + 2)
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(width, 50)

    try:
        wb.save(TRACKER_EXCEL)
        log.info("[monitor] Excel tracker saved: %s  (%d contracts)", TRACKER_EXCEL, len(sorted_rows))

        # Write directly to OneDrive synced folder (local) if path is set
        if ONEDRIVE_EXCEL_PATH:
            try:
                wb.save(ONEDRIVE_EXCEL_PATH)
                log.info("[monitor] Excel written directly to OneDrive folder: %s", ONEDRIVE_EXCEL_PATH)
            except PermissionError:
                log.warning("[monitor] OneDrive Excel is open — could not write directly, falling back to Power Automate.")
                _post_excel_to_onedrive(TRACKER_EXCEL)
        else:
            _post_excel_to_onedrive(TRACKER_EXCEL)

    except PermissionError:
        log.warning("[monitor] Excel tracker is open in another application — could not save.")
    except Exception as e:
        log.error("[monitor] Could not save Excel tracker: %s", e)


# ---------------------------------------------------------------------------
# OneDrive sync via Power Automate
# ---------------------------------------------------------------------------

def _post_excel_to_onedrive(excel_path: str) -> None:
    import time as _time
    webhook_url = os.environ.get("ONEDRIVE_EXCEL_WEBHOOK_URL", "")
    if not webhook_url:
        log.warning("[monitor] ONEDRIVE_EXCEL_WEBHOOK_URL not set — skipping OneDrive sync")
        return
    with open(excel_path, "rb") as f:
        content = base64.b64encode(f.read()).decode("utf-8")
    for attempt in range(1, 4):
        try:
            resp = requests.post(webhook_url, json={"content": content}, timeout=60)
            log.info("[monitor] Excel posted to OneDrive via Power Automate (HTTP %s)", resp.status_code)
            return
        except Exception as e:
            log.warning("[monitor] OneDrive post attempt %d/3 failed: %s", attempt, e)
            if attempt < 3:
                _time.sleep(10)
    log.error("[monitor] Failed to post Excel to OneDrive after 3 attempts")


# ---------------------------------------------------------------------------
# CSV exports (legacy — kept for reference, no longer called)
# ---------------------------------------------------------------------------

def _save_raw_all_columns_csv(headers: list, rows: list) -> None:
    """
    Save every column exactly as rendered on the Salesforce website (debug only).
    Writes to contracts_raw_debug.csv — does NOT overwrite the user-facing CSV.
    """
    debug_csv = os.path.join(DEBUG_DIR, "contracts_raw_debug.csv")
    try:
        # Remove empty trailing headers; pad / truncate each row to match
        clean_headers = [h if h else f"col_{i}" for i, h in enumerate(headers)]
        n = len(clean_headers)
        with open(debug_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(clean_headers)
            for row in rows:
                # Pad short rows, truncate long ones
                padded = (row + [""] * n)[:n]
                writer.writerow(padded)
        log.info("[monitor] Raw debug CSV saved: %s (%d rows, %d cols)",
                 debug_csv, len(rows), n)
    except Exception as e:
        log.error("[monitor] Could not save raw debug CSV: %s", e)


def _save_scraped_csv(contracts: dict) -> None:
    """
    Save structured scraped contracts (contract_no, account_name, status, created_date).
    No URL column — overwrites every run.
    """
    try:
        with open(SCRAPED_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["contract_no", "account_name", "status", "created_date"],
                                    extrasaction="ignore")
            writer.writeheader()
            for no, data in sorted(contracts.items(), key=lambda x: x[1].get("created_date", ""), reverse=True):
                writer.writerow({
                    "contract_no":  no,
                    "account_name": data.get("account_name", ""),
                    "status":       data.get("status", ""),
                    "created_date": data.get("created_date", ""),
                })
        log.info("[monitor] Scraped CSV updated: %s (%d rows)", SCRAPED_CSV, len(contracts))
    except Exception as e:
        log.error("[monitor] Could not save scraped CSV: %s", e)


def _save_approval_csv(in_approval: dict) -> None:
    """
    Save contracts currently in 'In Approval Process' to contracts_approval_master.csv.
    No URL column — overwrites every run.
    """
    try:
        with open(APPROVAL_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["contract_no", "account_name", "status", "created_date"],
                                    extrasaction="ignore")
            writer.writeheader()
            for no, data in sorted(in_approval.items(), key=lambda x: x[1].get("created_date", ""), reverse=True):
                writer.writerow({
                    "contract_no":  no,
                    "account_name": data.get("account_name", ""),
                    "status":       data.get("status", ""),
                    "created_date": data.get("created_date", ""),
                })
        log.info("[monitor] Approval master CSV updated: %s (%d rows)", APPROVAL_CSV, len(in_approval))
    except Exception as e:
        log.error("[monitor] Could not save approval CSV: %s", e)


# ---------------------------------------------------------------------------
# Master file persistence
# ---------------------------------------------------------------------------

def _load_master() -> dict:
    if not os.path.exists(MASTER_FILE):
        return {}
    try:
        with open(MASTER_FILE, "r") as f:
            return json.load(f)
    except Exception as e:
        log.warning("[monitor] Could not load master file: %s", e)
        return {}


def _save_master(master: dict) -> None:
    try:
        with open(MASTER_FILE, "w") as f:
            json.dump(master, f, indent=2)
    except Exception as e:
        log.error("[monitor] Could not save master file: %s", e)


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

def _parse_date(date_str: str):
    """
    Parse Salesforce date strings into a datetime object.
    Handles formats like:
        "08/07/2024, 3:14 pm"   → DD/MM/YYYY
        "2024-07-08"            → YYYY-MM-DD
        "07/08/2024"            → MM/DD/YYYY (fallback)
    Returns None if unparseable.
    """
    if not date_str:
        return None
    # Take only the date part (before comma)
    date_part = date_str.split(",")[0].strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_part, fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def _screenshot(name: str) -> None:
    try:
        _page.screenshot(path=os.path.join(DEBUG_DIR, f"{name}.png"))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Standalone entry point — spawned as a subprocess by webhook_listener.py
# Must be at end of file so all helpers above are defined before __main__ runs
# ---------------------------------------------------------------------------

def _build_teams_card(change: dict) -> dict:
    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.4",
        "body": [
            {
                "type": "TextBlock",
                "text": "⚡ Contract Status Changed",
                "weight": "Bolder",
                "size": "Medium",
                "color": "Accent",
            },
            {
                "type": "TextBlock",
                "text": change.get("account_name", "—"),
                "wrap": True,
                "spacing": "Small",
            },
            {
                "type": "FactSet",
                "spacing": "Medium",
                "facts": [
                    {"title": "Contract No.",    "value": change.get("contract_no",  "—")},
                    {"title": "Created Date",    "value": change.get("created_date", "—")},
                    {"title": "Previous Status", "value": change.get("old_status",   "—")},
                    {"title": "New Status",      "value": change.get("new_status",   "—")},
                ],
            },
        ],
        "actions": [
            {
                "type": "Action.OpenUrl",
                "title": "View Contract",
                "url": change.get("url", ""),
                "style": "positive",
            }
        ],
    }


if __name__ == "__main__":
    import time

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    POLL_INTERVAL = 15 * 60
    RETRY_DELAY   =  5 * 60

    while True:
        try:
            initialize_session()
            break
        except Exception as exc:
            log.error("[monitor] Session init failed: %s — retrying in %d min",
                      exc, RETRY_DELAY // 60, exc_info=True)
            time.sleep(RETRY_DELAY)

    while True:
        try:
            changes = check_for_changes()
            webhook_url = os.environ.get("TEAMS_CONTRACT_STATUS_WEBHOOK_URL", "")
            for change in changes:
                try:
                    card = _build_teams_card(change)
                    if not webhook_url:
                        log.error("[monitor] TEAMS_CONTRACT_STATUS_WEBHOOK_URL not set")
                        continue
                    resp = requests.post(webhook_url, json={"adaptive_card": card}, timeout=10)
                    log.info("[monitor] Posted card for %s (HTTP %s)",
                             change["contract_no"], resp.status_code)
                except Exception as exc:
                    log.error("[monitor] Failed to post card for %s: %s",
                              change.get("contract_no"), exc)
        except Exception as exc:
            log.error("[monitor] Poll error: %s", exc, exc_info=True)

        log.info("[monitor] Sleeping %d minutes.", POLL_INTERVAL // 60)
        time.sleep(POLL_INTERVAL)
